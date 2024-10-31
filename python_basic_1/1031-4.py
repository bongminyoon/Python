import openpyxl
import openpyxl.utils.cell as ut_cell
from openpyxl.styles import Alignment, Font
import os

wb = openpyxl.Workbook()
ws1 = wb.create_sheet(index=0, title='column')
print(wb.sheetnames)
wb.remove(wb['Sheet'])
print(wb.sheetnames)
ws2 = wb.create_sheet(index=1, title='Row')
print(wb.sheetnames)

for col in ws1.iter_cols(min_row=1, min_col=1, max_row=3, max_col=6):
    print(col)
    for each_cell in col:
        each_cell.value = ut_cell.get_column_letter(each_cell.column)
        each_cell.alignment = Alignment(horizontal='right', vertical='center')
        each_cell.font = Font(italic=True, name='consoras', size=10, color='ff0000')


wb.save(os.path.join(os.getcwd(), "output", "create_workbook2.xlsx"))
