import openpyxl
from openpyxl.styles import Alignment
import os

from openpyxl.styles.alignment import horizontal_alignments

wb = openpyxl.Workbook()

ws = wb.create_sheet(index=0, title='Merge')
wb.remove(wb['Sheet'])

tuple_of_rows = ((1,2),
                 (3,4),
                 (5,6),
                 (7,8),
                 (9,10),
                 )
for row in tuple_of_rows:
    ws.append(row)
    print(row)

ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
A1 = ws.cell(row=1, column=1)
A1.value = "Merged"
A1.alignment = Alignment(horizontal='center', vertical='center')

wb.save(os.path.join(os.getcwd(), "output", "create_workbook3.xlsx"))